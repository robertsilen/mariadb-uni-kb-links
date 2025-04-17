import mariadb
from openai import OpenAI
import json
import os
import re
import pandas as pd
from openpyxl import load_workbook
conn = mariadb.connect(
       host="127.0.0.1",
       port=3306,
       user="root",
       password="Password123!"
   )
cur = conn.cursor()
cur.execute("USE kb_content")
client = OpenAI(api_key=os.getenv('OPENAI_API_KEY_KB_UNI'))

def read_mariadb_markdown_files(directory='mariadb-for-universities'):
    files_data = []
    files = []
    for filename in os.listdir(directory):
        if filename.startswith('mariadb-') and filename.endswith('.md'):
            files.append(filename)
    files.sort()
    for filename in files:
        filepath = os.path.join(directory, filename)
        with open(filepath, 'r', encoding='utf-8') as f:
            content = f.read()
            files_data.append((filename, content))
    return files_data

def remove_front_matter(markdown_text):
    return markdown_text.split('---\n')[2]

def split_by_h3(markdown_text):
    segments = []
    current_segment = []
    current_title = None
    has_started = False

    for line in markdown_text.splitlines():
        if line.startswith('### '):
            if current_segment and current_title:
                segments.append({
                    'title': current_title,
                    'content': '\n'.join(current_segment)
                })
                current_segment = []
            current_title = line[4:].strip()  # Remove '### ' prefix
            has_started = True
        if has_started:
            current_segment.append(line)
    
    if current_segment and current_title:
        segments.append({
            'title': current_title,
            'content': '\n'.join(current_segment)
        })
    
    return segments


def test_search_kb(input_text="Installing MariaDB with package manager", n=3):
    closest_content = search_for_closest_content(input_text, n)
    print(input_text)
    for closest in closest_content:
        print(closest['distance'], closest['title'])

def embed(text):
    response = client.embeddings.create(
        input = text,
        model = "text-embedding-3-small" # max 8192 tokens (roughly 32k chars)
    )
    return response.data[0].embedding

def search_for_closest_content(text, n):
   embedding = embed(text)  # using same embedding model as in preparations
   cur.execute("""
       SELECT title, url, content,
              VEC_DISTANCE_EUCLIDEAN(embedding, VEC_FromText(%s)) AS distance
       FROM kb_content.kb_chunks
       ORDER BY distance ASC
       LIMIT %s;
   """, (str(embedding), n))

   closest_content = [
       {"title": title, "url": url, "content": content, "distance": distance}
       for title, url, content, distance in cur
   ]
   return closest_content

results = []
files_data = read_mariadb_markdown_files()
for name, text in files_data:
    text = remove_front_matter(text)
    segments = split_by_h3(text)
    print(f"File: {name}, Segments found: {len(segments)}")
    for segment in segments:
        segment['name'] = name
        print(segment['title'])
        closest_content = search_for_closest_content(segment['content'], 3)
        for i, closest in enumerate(closest_content):
            print(closest['distance'], closest['title'])
            segment[f'kb{i+1}_distance'] = closest['distance']
            segment[f'kb{i+1}_title'] = closest['title']
            segment[f'kb{i+1}_url'] = closest['url']
        results.append(segment)  # Append the entire segment dictionary instead of extending

df = pd.DataFrame(results)
df = df[['name'] + [col for col in df.columns if col != 'name']]

# Save to Excel
df.to_excel('results.xlsx', index=False)

# Adjust row heights
wb = load_workbook('results.xlsx')
ws = wb.active
for row in ws.iter_rows():
    ws.row_dimensions[row[0].row].height = 15
wb.save('02_results.xlsx')
